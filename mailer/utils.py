# mailer/utils.py
import csv
from typing import Dict, List, Tuple

from django.template import Template, Context

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore


def read_contacts(file_path: str) -> Tuple[List[Dict[str, str]], List[str]]:
    """
    Read .xlsx or .csv into list of row dicts and return (rows, headers).
    Each row dict keys equal the original header strings.
    """
    lower = file_path.lower()
    if lower.endswith(".xlsx"):
        if load_workbook is None:
            raise RuntimeError("openpyxl not installed")
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data = []
        for r in rows[1:]:
            item = {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r)}
            data.append(item)
        return data, headers
    elif lower.endswith(".csv"):
        with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            data = [{k: (v or "") for k, v in row.items()} for row in reader]
            headers = list(reader.fieldnames or [])
            return data, headers
    else:
        raise ValueError("Unsupported file type")


def render_template_string(template_str: str, variables: Dict[str, str]) -> str:
    django_template = Template(template_str)
    return django_template.render(Context(variables))
